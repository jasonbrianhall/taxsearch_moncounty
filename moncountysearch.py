#!/usr/bin/env python3
"""
Monongalia County Tax Search Tool
A command-line tool for searching tax records from the Monongalia County Sheriff's Office
"""

import requests
import argparse
import urllib3
import json
import logging
import os
import re
import time
import sys
from datetime import datetime

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Set up logging
def setup_logging(debug_level):
    log_levels = {
        0: logging.WARNING,
        1: logging.INFO,
        2: logging.DEBUG
    }
    level = log_levels.get(debug_level, logging.DEBUG)
    
    # Create logs directory if it doesn't exist
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # Set up file handler with timestamped filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = f'logs/tax_search_{timestamp}.log'
    
    # Configure logging
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    
    return logging.getLogger('tax_search')


def initialize_session(domain, logger, url=None):
    """Initialize a session with required cookies"""
    session = requests.Session()
    session.verify = False  # Disable SSL certificate verification
    
    # Set the initial cookies based on the website requirements
    cookies = {
        'DIST': '',
        'EOF': '1',
        'INDEXPGMF': 'INDEXF.html',
        'INDEXPGMX': 'INDEXX.html',
        'INDEXPGM': 'INDEX.html',
        'LYEAR': '0',
        'MAP': '',
        'PARC': '',
        'PMTINDEXF': 'INDEXFp.html',
        'PMTINDEXX': 'INDEXXp.html',
        'PMTINDEX': 'INDEXp.html',
        'PMTSEARCHF': 'SEARCHFp.html',
        'PMTSEARCHX': 'SEARCHXp.html',
        'PMTSEARCH': 'SEARCHp.html',
        'PMTTICKETF': 'TICKETFp.html',
        'PMTTICKETX': 'TICKETXp.html',
        'PMTTICKET': 'TICKETp.html',
        'PUB': 'B',
        'RECS': '33',
        'RN': '34',
        'RPB': 'B',
        'SEARCHPGMF': 'SEARCHF.html',
        'SEARCHPGMX': 'SEARCHX.html',
        'SEARCHPGM': 'SEARCH.html',
        'SEARCH': '1',
        'SPAGE': '1',
        'SPAR': '',
        'TICKETPGMF': 'TICKETF.html',
        'TICKETPGMX': 'TICKETX.html',
        'TICKETPGM': 'TICKET.html',
        'TPACCT': '',
        'TPNAME': '',
        'TPSX': '',
        'TPTICK': '',
        'TPTYR': ''
    }
    
    # Add the cookies to the session
    for key, value in cookies.items():
        session.cookies.set(key, value, domain=domain, path='/')
    
    logger.info("Session initialized with required cookies")
    logger.debug(f"Cookies set: {json.dumps(dict(session.cookies.items()), indent=2)}")
    
    return session

def search_by_name(session, name, common_params, domain, logger, max_pages=None, url=None):
    """Search tax records by taxpayer name (last first)"""
    logger.info(f"Performing name search for: {name}")
    
    # Update the TPNAME cookie with the search name
    session.cookies.set('TPNAME', name, domain=domain, path='/')
    
    # Start with the name-specific payload
    payload = {
        'TPNAME': name,
        'SBYNAME': 'Search by Name'
    }
    
    # Add common parameters to session cookies and payload
    apply_common_params(session, common_params, domain, logger)
    
    # Perform the search and handle pagination
    return perform_search_with_pagination(session, payload, domain, logger, max_pages, url)

def search_by_account(session, account_number, common_params, domain, logger, max_pages=None, url=None):
    """Search tax records by account number"""
    logger.info(f"Performing account search for: {account_number}")
    
    # Update the TPACCT cookie
    session.cookies.set('TPACCT', account_number, domain=domain, path='/')
    
    # Start with the account-specific payload
    payload = {
        'TPACCT': account_number,
        'SBYACCT': 'Search by Account'
    }
    
    # Add common parameters to session cookies and payload
    apply_common_params(session, common_params, domain, logger)
    
    # Perform the search and handle pagination
    return perform_search_with_pagination(session, payload, domain, logger, max_pages, url)

def search_by_ticket(session, year, ticket_number, suffix, common_params, domain, logger, max_pages=None, url=None):
    """Search tax records by tax year and ticket number"""
    logger.info(f"Performing ticket search for year: {year}, ticket: {ticket_number}, suffix: {suffix}")
    
    # Update relevant cookies
    session.cookies.set('TPTYR', year, domain=domain, path='/')
    session.cookies.set('TPTICK', ticket_number, domain=domain, path='/')
    session.cookies.set('TPSX', suffix, domain=domain, path='/')
    
    # Start with the ticket-specific payload
    payload = {
        'TPTYR': year,
        'TPTICK': ticket_number,
        'TPSX': suffix,
        'SBYTICKET': 'Search by Ticket'
    }
    
    # Add common parameters to session cookies and payload
    apply_common_params(session, common_params, domain, logger)
    
    # Perform the search and handle pagination
    return perform_search_with_pagination(session, payload, domain, logger, max_pages, url)

def search_by_map(session, district, map_num, parcel, sub_parcel, common_params, domain, logger, max_pages=None, url=None):
    """Search tax records by district, map, parcel, and sub-parcel"""
    logger.info(f"Performing map search with district: {district}, map: {map_num}, "
                f"parcel: {parcel}, sub-parcel: {sub_parcel}")
    
    # Update map-specific cookies
    session.cookies.set('DIST', district, domain=domain, path='/')
    session.cookies.set('MAP', map_num, domain=domain, path='/')
    session.cookies.set('PARC', parcel, domain=domain, path='/')
    session.cookies.set('SPAR', sub_parcel, domain=domain, path='/')
    
    # Start with the map-specific payload
    payload = {
        'SDIST': district,
        'SMAP': map_num,
        'SPAR': parcel,
        'SSPAR': sub_parcel,
        'SBYMAP': 'Search by Map/Parcel'
    }
    
    # Add common parameters to session cookies and payload
    apply_common_params(session, common_params, domain, logger)
    
    # Perform the search and handle pagination
    return perform_search_with_pagination(session, payload, domain, logger, max_pages, url)
    
def apply_common_params(session, params, domain, logger):
    """Apply common search parameters to the session cookies and payload"""
    # Map parameter names to cookie names
    param_to_cookie = {
        'limit_year': 'LYEAR',
        'prop_type': 'RPB',
        'status': 'PUB',
        'district': 'DIST',
    }
    
    if params:
        logger.info(f"Applying common search parameters: {params}")
        
        for param, value in params.items():
            if value and param in param_to_cookie:
                cookie_name = param_to_cookie[param]
                logger.debug(f"Setting {cookie_name} cookie to '{value}'")
                session.cookies.set(cookie_name, value, domain=domain, path='/')
                
                # Also add to the URL payload if needed
                if param == 'limit_year':
                    session.cookies.set('lyear', value, domain=domain, path='/')
                elif param == 'prop_type':
                    session.cookies.set('rpb', value, domain=domain, path='/')
                elif param == 'status':
                    session.cookies.set('pub', value, domain=domain, path='/')
                elif param == 'district':
                    # Also set SDIST for forms that use it
                    session.cookies.set('SDIST', value, domain=domain, path='/')

def perform_search_with_pagination(session, initial_payload, domain, logger, max_pages=None, url=None):
    """Perform a search and handle pagination"""
    logger.info(f"Starting search with pagination (max_pages={max_pages})")
    
    # Initialize variables to track results and pagination
    all_results = []
    current_page = 1
    total_pages = None
    headers = ["Ticket", "Type", "Taxpayer Name", "Address", "Half Yr Tax", "Page"]
    
    # Use tqdm for progress bar if available
    try:
        from tqdm import tqdm
        progress_enabled = True
    except ImportError:
        progress_enabled = False
        logger.warning("tqdm not installed, progress bar disabled")
    
    # Perform the initial search to get first page of results
    logger.info(f"Processing page {current_page}")
    print(f"Processing page {current_page}...", end="\r", flush=True)
    
    # Set the page cookie
    session.cookies.set('SPAGE', str(current_page), domain=domain, path='/')
    
    # Get first page of results
    response = perform_search(session, initial_payload, domain, logger, url)
    if not response:
        logger.error("Initial search failed")
        return "Search failed - no response from server"
    
    # Extract page information
    total_pages, page_results = extract_search_results(response, logger, current_page)
    
    if not page_results:
        logger.warning("No results found on first page")
        return "No tax records found"
    
    # Add first page results to our collection
    all_results.extend(page_results)
    
    # Create progress bar if more pages available and tqdm is installed
    if total_pages and total_pages > 1 and progress_enabled:
        # Calculate remaining pages (we already did page 1)
        remaining_pages = total_pages - 1
        if max_pages:
            remaining_pages = min(remaining_pages, max_pages - 1)
        
        # Create progress bar
        progress_bar = tqdm(total=remaining_pages, desc="Fetching pages", unit="page")
    else:
        progress_bar = None
    
    # Process all remaining pages
    while (total_pages is not None and current_page < total_pages) and (max_pages is None or current_page < max_pages):
        # Move to next page
        current_page += 1
        
        logger.info(f"Processing page {current_page} of {total_pages}")
        print(f"Processing page {current_page} of {total_pages}...", end="\r", flush=True)
        
        # Update progress bar
        if progress_bar:
            progress_bar.update(1)
        
        # For pagination, we need to use the NEXT link/button which requires a special payload
        # The HTML shows 'SEARCH.html?TASK=NEXT' is used for the Next page link
        next_payload = {'TASK': 'NEXT'}
        
        # Get the next page of results
        response = perform_search(session, next_payload, domain, logger, url)
        if not response:
            logger.error(f"Failed to get page {current_page}")
            break
        
        # Extract results from this page
        _, page_results = extract_search_results(response, logger, current_page)
        
        if not page_results:
            logger.warning(f"No results found on page {current_page}")
            break
        
        # Add this page's results to our collection
        all_results.extend(page_results)
    
    # Close progress bar if it was created
    if progress_bar:
        progress_bar.close()
    
    # Final count info
    logger.info(f"Retrieved {len(all_results)} records from {current_page} page(s)")
    
    # Return the combined results with headers
    return {"headers": headers, "data": all_results, "pagination": {"current_page": current_page, "total_pages": total_pages}}

def perform_search(session, payload, domain, logger, url=None):
    """Execute the search with the provided payload"""
    # Use the provided URL if available, otherwise construct using domain
    if url:
        if "SEARCH.html" not in url:
            # If URL doesn't end with SEARCH.html, we'll append it
            if url.endswith('/'):
                url = f"{url}SEARCH.html"
            else:
                url = f"{url}/SEARCH.html"
    else:
        url = f"https://{domain}/SEARCH.html"
    
    logger.debug(f"Request URL: {url}")
    logger.debug(f"Request Payload: {json.dumps(payload, indent=2)}")
    logger.debug(f"Request Cookies: {json.dumps(dict(session.cookies.items()), indent=2)}")
    
    try:
        start_time = time.time()
        response = session.post(url, data=payload, verify=False)
        elapsed_time = time.time() - start_time
        
        logger.info(f"Response received in {elapsed_time:.2f} seconds with status code: {response.status_code}")
        
        if response.status_code != 200:
            logger.error(f"Error: Received status code {response.status_code}")
            return None
        
        # Save the raw HTML response to a file for inspection
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response_file = f'logs/response_{timestamp}.html'
        with open(response_file, 'w', encoding='utf-8') as f:
            f.write(response.text)
        logger.info(f"Raw response saved to: {response_file}")
        
        return response.text
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Request failed: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return None

def extract_search_results(html_content, logger, current_page):
    """Extract search results from the HTML response"""
    # Find the total pages info - looking for the format in the example: "Page     1 of &nbsp; 1232"
    total_pages = None
    page_info_match = re.search(r'Page\s+\d+\s+of\s+&nbsp;\s*(\d+)', html_content)
    if page_info_match:
        total_pages = int(page_info_match.group(1))
        logger.info(f"Total pages: {total_pages}")
    else:
        # Try a more generic pattern as fallback
        page_info_match = re.search(r'Page\s+\d+\s+of\s+(\d+)', html_content)
        if page_info_match:
            total_pages = int(page_info_match.group(1))
            logger.info(f"Total pages: {total_pages}")
    
    # Extract each record from the table rows
    results = []
    
    # Find all table rows with tax record data - pattern based on the example HTML
    row_pattern = r'<tr class="[^"]*">\s*<TD class=left[^>]*>.*?<A href="TICKET\.html\?TPTYR=(\d+)&amp;TPTICK=(\d+)&amp;TPSX=([^"]*)"[^>]*>(\d+ -\s*\d+\s*[^<]*)</a>.*?</TD>\s*<td>.*?(?:<A[^>]*>([^<]*)</A>|([^<]*)).*?</td>\s*<td>.*?<font class="tdtext">([^<]*)</font></td>\s*<td>.*?<font class="tdtext">([^<]*)</font></td>\s*<td[^>]*>.*?<div[^>]*>.*?<div[^>]*>.*?</div>\s*([^<]*)</div>\s*</td>\s*</tr>'
    
    record_matches = re.finditer(row_pattern, html_content, re.DOTALL)
    
    for match in record_matches:
        tax_year = match.group(1)
        ticket_number = match.group(2)
        suffix = match.group(3).strip()
        ticket_display = match.group(4).strip()
        
        # Type might be in either group 5 or 6 depending on if there's a link
        record_type = match.group(5) if match.group(5) else match.group(6)
        record_type = record_type.strip() if record_type else ""
        
        taxpayer_name = match.group(7).strip()
        address = match.group(8).strip()
        amount = match.group(9).strip()
        
        # Create a data row and add page number
        row = [ticket_display, record_type, taxpayer_name, address, amount, f"Page {current_page}"]
        results.append(row)
    
    logger.info(f"Extracted {len(results)} records from page {current_page}")
    return total_pages, results

def save_results_to_file(results, output_file, logger):
    """Save results to a file (CSV, JSON, Excel, or text)"""
    logger.info(f"Saving results to {output_file}")
    
    # Create the directory if it doesn't exist
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Determine file type from extension
    _, ext = os.path.splitext(output_file)
    ext = ext.lower()
    
    try:
        if ext == '.xlsx':
            save_to_excel(results, output_file, logger)
        elif ext == '.json':
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)
        elif ext == '.csv':
            import csv
            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Write headers
                if isinstance(results, dict) and 'headers' in results and 'data' in results:
                    writer.writerow(results['headers'])
                    writer.writerows(results['data'])
                elif isinstance(results, list):
                    writer.writerows(results)
                else:
                    # Error message
                    writer.writerow([results])
        else:
            # Default to text format
            with open(output_file, 'w', encoding='utf-8') as f:
                if isinstance(results, dict) and 'headers' in results and 'data' in results:
                    f.write('\t'.join(results['headers']) + '\n')
                    f.write('-' * 80 + '\n')
                    for row in results['data']:
                        f.write('\t'.join(str(cell) for cell in row) + '\n')
                elif isinstance(results, list):
                    for row in results:
                        f.write('\t'.join(str(cell) for cell in row) + '\n')
                else:
                    f.write(str(results))
        
        logger.info(f"Results successfully saved to {output_file}")
        return True
    except Exception as e:
        logger.error(f"Failed to save results: {str(e)}")
        return False

def save_to_excel(results, output_file, logger):
    """Save results to an Excel file with formatting"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl is not installed. Please install it with: pip install openpyxl")
        print("Error: openpyxl is not installed. Please install it with: pip install openpyxl")
        return False
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Records"
    
    # Format and populate the data
    if isinstance(results, dict) and 'headers' in results and 'data' in results:
        # Prepare styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        center_alignment = Alignment(horizontal='center')
        
        # Write headers
        for col_idx, header in enumerate(results['headers'], 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Write data rows
        for row_idx, row_data in enumerate(results['data'], 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(results['headers'], 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = max(12, len(header) + 2)
    
    # Create a second sheet with search metadata
    meta_sheet = wb.create_sheet(title="Search Info")
    
    # Add search metadata
    meta_sheet['A1'] = "Search Date"
    meta_sheet['B1'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Add record counts
    meta_sheet['A2'] = "Records Found"
    if isinstance(results, dict) and 'data' in results:
        meta_sheet['B2'] = len(results['data'])
        
        if 'pagination' in results:
            meta_sheet['A3'] = "Current Page"
            meta_sheet['B3'] = results['pagination'].get('current_page', 1)
            meta_sheet['A4'] = "Total Pages"
            meta_sheet['B4'] = results['pagination'].get('total_pages', 1)
    
    # Format metadata
    for row in range(1, 5):
        meta_sheet.cell(row=row, column=1).font = Font(bold=True)
    
    # Set column widths for metadata
    meta_sheet.column_dimensions['A'].width = 15
    meta_sheet.column_dimensions['B'].width = 25
    
    # Save the workbook
    wb.save(output_file)
    return True

def inspect_log_file(filename=None):
    """Utility function to inspect the most recent log file"""
    # Find the most recent log file if not specified
    if not filename:
        log_dir = 'logs'
        if not os.path.exists(log_dir):
            print("No logs directory found.")
            return
            
        html_files = [f for f in os.listdir(log_dir) if f.startswith('response_') and f.endswith('.html')]
        if not html_files:
            print("No response log files found.")
            return
            
        # Sort by modification time (newest first)
        html_files.sort(key=lambda f: os.path.getmtime(os.path.join(log_dir, f)), reverse=True)
        filename = os.path.join(log_dir, html_files[0])
    
    # Open and read the file
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Print basic info
        print(f"File: {filename}")
        print(f"Size: {len(content)} bytes")
        
        # Look for page info
        page_info = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', content)
        if page_info:
            print(f"Page Information: Page {page_info.group(1)} of {page_info.group(2)}")
        
        # Count records
        row_pattern = r'<tr class="[^"]*">\s*<TD class=left[^>]*>.*?<A href="TICKET\.html\?TPTYR=(\d+)&amp;TPTICK=(\d+)&amp;TPSX=([^"]*)"'
        records = re.findall(row_pattern, content)
        print(f"Number of records found: {len(records)}")
        
        # Show sample records if any found
        if records:
            print("\nSample Records (first 3):")
            sample_pattern = r'<tr class="[^"]*">\s*<TD class=left[^>]*>.*?<A[^>]*>(.*?)</a>.*?</TD>\s*<td>.*?(?:<A[^>]*>([^<]*)</A>|([^<]*)).*?</td>\s*<td>.*?<font class="tdtext">([^<]*)</font></td>\s*<td>.*?<font class="tdtext">([^<]*)</font></td>\s*<td[^>]*>.*?<div[^>]*>.*?<div[^>]*>.*?</div>\s*([^<]*)</div>'
            sample_records = re.findall(sample_pattern, content, re.DOTALL)
            
            for i, record in enumerate(sample_records[:3]):
                ticket = record[0].strip() if record[0] else ""
                record_type = (record[1] or record[2]).strip()
                name = record[3].strip()
                address = record[4].strip()
                amount = record[5].strip()
                print(f"{i+1}. Ticket: {ticket}, Type: {record_type}, Name: {name}, Address: {address}, Amount: {amount}")
        
        print("\nFile inspection complete")
        
    except Exception as e:
        print(f"Error inspecting file: {str(e)}")

def main():
    """Main function to run the tax search tool"""
    # Create the main parser
    parser = argparse.ArgumentParser(description='County Tax Search Tool')
    
    # Add common arguments to the main parser
    parser.add_argument('--domain', '-dm', default='monongalia.softwaresystems.com',
                      help='Domain to search (default: monongalia.softwaresystems.com)')
    parser.add_argument('--url', '-u', default='https://monongalia.softwaresystems.com/SEARCH.html',
                      help='Full URL to search (default: https://monongalia.softwaresystems.com/SEARCH.html)')
    parser.add_argument('--output', '-o', help='Output file for results (supports .xlsx, .json, .csv, .txt)')
    parser.add_argument('--debug', '-v', action='count', default=0, 
                      help='Increase verbosity (use -v, -vv for more detailed logging)')
    parser.add_argument('--timeout', '-to', type=int, default=30, 
                      help='Request timeout in seconds (default: 30)')
    parser.add_argument('--user-agent', '-ua', default='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                      help='User agent string to use for requests')
    parser.add_argument('--max-pages', '-mp', type=int, default=None,
                      help='Maximum number of pages to retrieve (default: all)')
    parser.add_argument('--inspect', '-i', nargs='?', const='', 
                      help='Inspect a log file (or most recent if not specified)')
    
    # Add common filtering arguments to the main parser
    parser.add_argument('--limit-year', '-ly', help='Limit search to a specific tax year')
    parser.add_argument('--prop-type', '-pt', choices=['B', 'R', 'P'], default='B', 
                      help='Property type: B=Both, R=Real, P=Personal')
    parser.add_argument('--status', '-st', choices=['B', 'P', 'U'], default='B',
                      help='Payment status: B=Both, P=Paid, U=Unpaid')
    parser.add_argument('--district', '-d', help='Filter by district code (e.g., 01, 02, etc.)')
    
    # Create subparsers for the different search types
    subparsers = parser.add_subparsers(dest='search_type', help='Type of search to perform')
    
    # Name search parser
    name_parser = subparsers.add_parser('name', help='Search by taxpayer name')
    name_parser.add_argument('name', help='Taxpayer name (last first)')
    
    # Account search parser
    acct_parser = subparsers.add_parser('account', help='Search by account number')
    acct_parser.add_argument('account', help='Taxpayer account number')
    
    # Ticket search parser
    ticket_parser = subparsers.add_parser('ticket', help='Search by ticket number')
    ticket_parser.add_argument('year', help='Tax year (4 digits)')
    ticket_parser.add_argument('ticket', help='Ticket number')
    ticket_parser.add_argument('--suffix', '-s', default='', help='Ticket suffix (optional)')
    
    # Map/Parcel search parser
    map_parser = subparsers.add_parser('map', help='Search by district/map/parcel')
    map_parser.add_argument('--district', '-d', default='', help='District code')
    map_parser.add_argument('--map', '-m', default='', help='Map number')
    map_parser.add_argument('--parcel', '-p', default='', help='Parcel number')
    map_parser.add_argument('--subparcel', '-sp', default='', help='Sub-parcel number')
    
    # Parse arguments
    args = parser.parse_args()
    
    # If inspect mode, just inspect the log file and exit
    if args.inspect is not None:
        inspect_log_file(args.inspect)
        return
    
    # Set up logging based on debug level
    logger = setup_logging(args.debug)
    logger.info(f"Tax Search Tool started with arguments: {vars(args)}")
    
    # If no search type was provided, show help and exit
    if not args.search_type:
        parser.print_help()
        logger.info("No search type specified, exiting.")
        return
    
    # Check for required dependencies if Excel output is requested
    if args.output and args.output.lower().endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError:
            print("Error: openpyxl is required for Excel output. Please install it with: pip install openpyxl")
            logger.error("openpyxl dependency not found for Excel output")
            return
    
    # Initialize session with cookies
    session = initialize_session(args.domain, logger)
    
    # Set custom user agent and timeout
    session.headers.update({'User-Agent': args.user_agent})
    session.timeout = args.timeout
    
    # Collect common parameters for all search types
    common_params = {
        'limit_year': args.limit_year,
        'prop_type': args.prop_type,
        'status': args.status,
        'district': args.district,  # Add district to common params
    }
    
    print(f"Starting search... (Type: {args.search_type}, URL: {args.url})")
    
    # Perform the appropriate search based on the arguments
    if args.search_type == 'name':
        results = search_by_name(session, args.name, common_params, args.domain, logger, args.max_pages, args.url)
    elif args.search_type == 'account':
        results = search_by_account(session, args.account, common_params, args.domain, logger, args.max_pages, args.url)
    elif args.search_type == 'ticket':
        results = search_by_ticket(session, args.year, args.ticket, args.suffix, common_params, args.domain, logger, args.max_pages, args.url)
    elif args.search_type == 'map':
        results = search_by_map(
            session, args.district, args.map, args.parcel, args.subparcel,
            common_params, args.domain, logger, args.max_pages, args.url
        )
    else:
        logger.error(f"Unknown search type: {args.search_type}")
        parser.print_help()
        return
    
    # Save results to a file if output file was specified
    if args.output and results:
        if save_results_to_file(results, args.output, logger):
            print(f"\nResults saved to {args.output}")
        else:
            print("\nFailed to save results. See log for details.")
    elif isinstance(results, dict) and 'data' in results:
        # Display summary of results
        print(f"\nFound {len(results['data'])} records")
        if results['pagination']['total_pages'] > 1:
            print(f"Showing page {results['pagination']['current_page']} of {results['pagination']['total_pages']}")
            if args.max_pages and results['pagination']['current_page'] < results['pagination']['total_pages']:
                print(f"Note: Only retrieved {args.max_pages} of {results['pagination']['total_pages']} total pages")
    else:
        # Just print the results (likely an error message)
        print(f"\n{results}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nSearch cancelled by user.")
    except Exception as e:
        print(f"\nError: {str(e)}")
        print("For more details, check the log files in the 'logs' directory.")
        
        # Try to get traceback for better debugging
        import traceback
        traceback.print_exc()
